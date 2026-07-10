#파일 불러오는 기능 함수 import
from openpyxl import load_workbook

#1. 파일불러와서 워크북 객체 만들기
wb= load_workbook('./01_excel/sample3.xlsx')
#2. 활성화된 sheet 참조
sheet= wb.active

#3. D1~D10 셀의 값들 읽어오기
for n in range(1,11):
    print(sheet.cell(row=n, column=4).value  , end=" ")
print()

#위 반복문은 값이 10개임을 알았기에 range(1,11)이라고 쓴 것임.. 몇 줄인지 모르면..?
for n in range(1, sheet.max_row+1):
    print(sheet.cell(row=n, column=4).value, end=" ")
